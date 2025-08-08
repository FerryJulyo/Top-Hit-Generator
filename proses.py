import os
import io
import msoffcrypto
import tkinter as tk
import re
import hashlib
import threading
import datetime
import pandas as pd
import xlrd
import tempfile

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from tempfile import NamedTemporaryFile
from xlrd.biffh import XLRDError
from tkinter import filedialog, messagebox, ttk, StringVar
from Crypto.Cipher import AES
from Crypto.Util.Padding import pad, unpad
from collections import defaultdict

class GenerateTopHit:
    def __init__(self, root):
        self.root = root
        self.root.title("Generate Top Hit")
        self.root.geometry("720x640")

        self.source_folder = tk.StringVar()
        self.destination_folder = tk.StringVar()
        self.database_file = tk.StringVar()
        self.encryption_key = tk.StringVar(value="60132323abcd")

        self.total_files = 0
        self.success_count = 0
        self.failure_count = 0

        self.df_db = None
        self.df_db_sing = None

        self.setup_ui()

    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main_frame, text="Folder Sumber:").grid(row=0, column=0, sticky=tk.W)
        ttk.Entry(main_frame, textvariable=self.source_folder, width=60).grid(row=0, column=1, padx=10)
        self.btn_browse_source = ttk.Button(main_frame, text="Pilih", command=self.browse_source)
        self.btn_browse_source.grid(row=0, column=2)

        ttk.Label(main_frame, text="Folder Tujuan:").grid(row=1, column=0, sticky=tk.W, pady=(10, 0))
        ttk.Entry(main_frame, textvariable=self.destination_folder, width=60).grid(row=1, column=1, padx=10, pady=(10, 0))
        self.btn_browse_destination = ttk.Button(main_frame, text="Pilih", command=self.browse_destination)
        self.btn_browse_destination.grid(row=1, column=2, pady=(10, 0))

        ttk.Label(main_frame, text="File Database (xlsx):").grid(row=2, column=0, sticky=tk.W, pady=(10, 0))
        ttk.Entry(main_frame, textvariable=self.database_file, width=60).grid(row=2, column=1, padx=10, pady=(10, 0))
        self.btn_browse_database = ttk.Button(main_frame, text="Pilih", command=self.browse_database)
        self.btn_browse_database.grid(row=2, column=2, pady=(10, 0))

        # self.btn_start = ttk.Button(main_frame, text="Mulai Proses", command=self.start_decryption)
        # self.btn_start.grid(row=3, column=0, columnspan=3, pady=20)

        self.progress = ttk.Progressbar(main_frame, orient="horizontal", length=400, mode="determinate")
        self.progress.grid(row=4, column=0, columnspan=3, pady=(0, 10))

        self.progress_label = ttk.Label(main_frame, text="")
        self.progress_label.grid(row=5, column=0, columnspan=3)

        progress_log_frame = ttk.LabelFrame(main_frame, text="Log Progres File", padding="10")
        progress_log_frame.grid(row=6, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        progress_log_frame.columnconfigure(0, weight=1)
        progress_log_frame.rowconfigure(0, weight=1)

        self.progress_text = tk.Text(progress_log_frame, height=10, wrap=tk.WORD)
        self.progress_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        scrollbar = ttk.Scrollbar(progress_log_frame, orient="vertical", command=self.progress_text.yview)
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.progress_text.configure(yscrollcommand=scrollbar.set)

        log_frame = ttk.LabelFrame(main_frame, text="Log Status", padding="10")
        log_frame.grid(row=7, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)

        self.log_text = tk.Text(log_frame, height=6, wrap=tk.WORD)
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.log_text.configure(yscrollcommand=scrollbar.set)

        self.vod_mode = tk.StringVar(value="VOD1")
        ttk.Label(main_frame, text="Mode:").grid(row=3, column=0, sticky=tk.W, pady=(10, 0))
        ttk.Radiobutton(main_frame, text="VOD1", variable=self.vod_mode, value="VOD1").grid(row=3, column=1, sticky=tk.W, pady=(10, 0))
        ttk.Radiobutton(main_frame, text="VOD2", variable=self.vod_mode, value="VOD2").grid(row=3, column=1, sticky=tk.E, pady=(10, 0))

        self.btn_start = ttk.Button(main_frame, text="Mulai Proses", command=self.start_decryption)
        self.btn_start.grid(row=4, column=0, columnspan=3, pady=20)

        # Adjust grid positions for remaining elements
        self.progress.grid(row=5, column=0, columnspan=3, pady=(0, 10))
        self.progress_label.grid(row=6, column=0, columnspan=3)
        progress_log_frame.grid(row=7, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        log_frame.grid(row=8, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)

    def browse_source(self):
        folder = filedialog.askdirectory()
        if folder:
            self.source_folder.set(folder)

    def browse_destination(self):
        folder = filedialog.askdirectory()
        if folder:
            self.destination_folder.set(folder)

    def browse_database(self):
        file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file:
            self.database_file.set(file)

    def log_message(self, message):
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def log_progress(self, message):
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        self.progress_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.progress_text.see(tk.END)
        self.root.update_idletasks()

    def start_decryption(self):
        self.set_buttons_state(tk.DISABLED)
        thread = threading.Thread(target=self.decrypt_all_files)
        thread.start()

    def set_buttons_state(self, state):
        self.btn_browse_source.config(state=state)
        self.btn_browse_destination.config(state=state)
        self.btn_browse_database.config(state=state)
        self.btn_start.config(state=state)


    def decrypt_all_files(self):
        src = self.source_folder.get()
        dst = self.destination_folder.get()
        db_path = self.database_file.get()
        vod_mode = self.vod_mode.get()

        # Validation checks
        if not os.path.isdir(src):
            messagebox.showerror("Error", "Folder sumber tidak valid.")
            return
        if not os.path.isdir(dst):
            messagebox.showerror("Error", "Folder tujuan tidak valid.")
            return
        if not os.path.isfile(db_path):
            messagebox.showerror("Error", "File database tidak ditemukan.")
            return

        try:
            self.log_message("Sedang membaca file master...")
            self.df_db = pd.read_excel(db_path, sheet_name="Song", usecols=["SongId", "Song","RomanSong","Collector", "Sing1","SingId1", "Sing2","SingId2", "Sing3","SingId3", "Sing4","SingId4", "Sing5","SingId5"],dtype=str)
            self.df_db_del = pd.read_excel(db_path, sheet_name="Delete Song", usecols=["SongId", "Song","RomanSong","Collector", "Sing1","SingId1", "Sing2","SingId2", "Sing3","SingId3", "Sing4","SingId4", "Sing5","SingId5"],dtype=str)
            self.df_db_sing = pd.read_excel(db_path, sheet_name="Sing", usecols=["SingId", "RomanSing"],dtype=str)
            # Step 1: Filter df_db_del dengan syarat SongId 7-9 karakter alfanumerik
            filtered_df_del = self.df_db_del[
                self.df_db_del["SongId"].apply(lambda x: isinstance(x, str) and x.isalnum() and 7 <= len(x) <= 9)
            ].copy()

            # Step 2: Cari baris yang SongId-nya belum ada di df_db
            missing_rows = filtered_df_del[~filtered_df_del["SongId"].isin(self.df_db["SongId"])]

            # Step 3: Gabungkan ke df_db
            self.df_db = pd.concat([self.df_db, missing_rows], ignore_index=True)

        except Exception as e:
            self.log_message(f"Gagal membaca database: {e}")
            return

        # Reset counters
        self.total_files = 0
        self.success_count = 0
        self.failure_count = 0
        self.progress_label.config(text="")
        self.progress_text.delete(1.0, tk.END)
        self.log_text.delete(1.0, tk.END)

        processed_files = []

        if vod_mode == "VOD2":
            # VOD2 Mode: Process TXT files only
            txt_files = [f for f in os.listdir(src) if f.endswith(".txt")]
            self.total_files = len(txt_files)

            if self.total_files == 0:
                self.log_message("Tidak ada file .txt ditemukan di folder sumber.")
                return

            for idx, file_name in enumerate(txt_files):
                txt_file = os.path.join(src, file_name)
                xls_file = os.path.join(dst, file_name[:-4] + ".xlsx")
                
                try:
                    # Convert TXT to XLS
                    if self.process_txt_to_xls(txt_file, xls_file):
                        processed_files.append(xls_file)
                        self.success_count += 1
                        self.log_message(f"✓ Berhasil: {os.path.basename(txt_file)}")
                        self.log_progress(f"✓ {os.path.basename(txt_file)} berhasil dikonversi ke XLSX")
                    else:
                        self.failure_count += 1
                        self.log_message(f"✗ Gagal: {os.path.basename(txt_file)}")
                except Exception as e:
                    self.failure_count += 1
                    self.log_message(f"✗ Gagal: {os.path.basename(txt_file)} - {str(e)}")
                    self.log_progress(f"✗ {os.path.basename(txt_file)} gagal: {str(e)}")

                # Update progress
                self.progress["value"] = idx + 1
                self.progress_label.config(text=f"{idx + 1}/{self.total_files} file diproses")
                self.root.update_idletasks()

        else:
            # VOD1 Mode: Process encrypted files only
            enc_files = [f for f in os.listdir(src) if f.endswith(".enc")]
            self.total_files = len(enc_files)

            if self.total_files == 0:
                self.log_message("Tidak ada file .enc ditemukan di folder sumber.")
                return

            for idx, file_name in enumerate(enc_files):
                enc_file = os.path.join(src, file_name)
                dec_file = os.path.join(dst, file_name[:-4] + ".xls")
                
                try:
                    # Decrypt the file
                    self.decrypt_file(enc_file, dec_file, self.encryption_key.get())
                    processed_files.append(dec_file)
                    self.success_count += 1
                    self.log_message(f"✓ Berhasil: {os.path.basename(enc_file)}")
                    self.log_progress(f"✓ {os.path.basename(enc_file)} berhasil didekripsi")
                except Exception as e:
                    self.failure_count += 1
                    self.log_message(f"✗ Gagal: {os.path.basename(enc_file)} - {str(e)}")
                    self.log_progress(f"✗ {os.path.basename(enc_file)} gagal: {str(e)}")

                # Update progress
                self.progress["value"] = idx + 1
                self.progress_label.config(text=f"{idx + 1}/{self.total_files} file diproses")
                self.root.update_idletasks()

        # Process and merge the files if we have any successful conversions
        if processed_files:
            total_export_steps = len(processed_files) + len(set(os.path.basename(f)[:5] for f in processed_files)) * 2 + 1
            self.progress["maximum"] = total_export_steps
            self.process_and_merge_data(processed_files, dst, progress_update_callback=self.update_progress)
            self.progress["value"] = self.progress["maximum"]

        # Show summary
        summary = f"Selesai! {self.success_count} berhasil, {self.failure_count} gagal dari {self.total_files} file."
        self.set_buttons_state(tk.NORMAL)
        self.log_message(summary)
        messagebox.showinfo("Selesai", summary)
    
    def decrypt_file(self, in_file, out_file, key):
        hashed_key = hashlib.sha256(key.encode()).digest()
        cipher = AES.new(hashed_key, AES.MODE_ECB)

        with open(in_file, "rb") as f:
            ciphertext = f.read()
        plaintext = unpad(cipher.decrypt(ciphertext), AES.block_size)

        with open(out_file, "wb") as f:
            f.write(plaintext)

    def process_and_merge_data(self, file_list, output_path, progress_update_callback=None):
        result = defaultdict(lambda: defaultdict(int))
        all_data = defaultdict(lambda: defaultdict(int))
        collector_result = defaultdict(lambda: defaultdict(int))
        collector_all_data = defaultdict(lambda: defaultdict(int))
        # 🆕 Rekap semua data

        # Buat dictionary lookup dari master
        song_dict = {}
        sing_dict = {}
        if self.df_db is not None:
            for _, row in self.df_db.iterrows():
                song_id = str(row["SongId"]).strip()
                song_name = str(row["Song"]).strip() if not pd.isna(row["Song"]) else ""
                rosong_name = str(row["RomanSong"]).strip() if not pd.isna(row["RomanSong"]) else ""
                label = str(row["Collector"]).strip() if not pd.isna(row["Collector"]) else ""
                singers = [str(row[col]).strip() for col in ["Sing1", "Sing2", "Sing3", "Sing4", "Sing5"] if not pd.isna(row[col]) and str(row[col]).strip()]
                singers_id = [str(row[col]).strip() for col in ["SingId1", "SingId2", "SingId3", "SingId4", "SingId5"] if not pd.isna(row[col]) and str(row[col]).strip()]
                song_dict[song_id] = {
                    "Song": song_name,
                    "RomanSong": rosong_name,
                    "Label": label,
                    "Singer": " - ".join(singers),
                    "SingId": " - ".join(singers_id)
                }

        if self.df_db_sing is not None:
            for _, row in self.df_db_sing.iterrows():
                sing_id = str(row["SingId"]).strip()
                roman_name = str(row["RomanSing"]).strip() if not pd.isna(row["RomanSing"]) else ""
                sing_dict[sing_id] = {
                    "RomanSing": roman_name
                }

        # Proses semua file terenkripsi
        for file in file_list:
            try:
                # 1. Verifikasi file exists
                if not os.path.exists(file):
                    self.log_message(f"File tidak ditemukan: {file}")
                    continue

                # 2. Coba baca langsung dengan pandas (untuk file tidak terenkripsi)
                try:
                    # 1. Coba baca langsung tanpa dekripsi
                    df = pd.read_excel(file, sheet_name="Lap1")
                    self.log_message(f"File {file} berhasil dibaca tanpa dekripsi")
                except:
                    # 2. Jika gagal, coba dekripsi dengan msoffcrypto
                    try:
                        with open(file, "rb") as f:
                            office_file = msoffcrypto.OfficeFile(f)
                            office_file.load_key(password="secret")  # Ganti dengan password yang sesuai
                            
                            # Decrypt file ke dalam memory
                            decrypted = io.BytesIO()
                            office_file.decrypt(decrypted)

                        # Simpan ke temporary file dari hasil decrypt
                        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                            temp_file.write(decrypted.getvalue())
                            temp_path = temp_file.name

                        try:
                            # 3. Coba baca dari hasil decrypt
                            df = pd.read_excel(temp_path, sheet_name="Lap1")
                            self.log_message(f"File {file} berhasil dibaca setelah dekripsi")
                        except Exception as e:
                            self.log_message(f"Gagal baca file setelah dekripsi: {str(e)}")
                            continue
                        finally:
                            # 4. Hapus file sementara
                            try:
                                os.unlink(temp_path)
                            except:
                                pass

                    except Exception as e:
                        self.log_message(f"Gagal proses file terenkripsi {file}: {str(e)}")
                        continue

                # 4. Proses data dari DataFrame
                group_key = os.path.basename(file)[:5]
                
                for index, row in df.iterrows():
                    try:
                        # Gunakan iloc untuk akses kolom by index
                        song_id = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                        jumlah_cell = row.iloc[1] if len(row) > 1 else None
                        
                        if not song_id or pd.isna(jumlah_cell):
                            continue
                            
                        try:
                            # Handle berbagai format angka
                            jumlah = int(float(str(jumlah_cell))) if str(jumlah_cell).strip() else 0
                            if jumlah == 0:
                                continue
                                
                            # Update hasil
                            result[group_key][song_id] += jumlah
                            all_data["ALL"][song_id] += jumlah

                            label = song_dict.get(song_id, {}).get("Label", "").upper()
                            collector_group = label if label in ["KCI", "WAMI", "RAI"] else "Lain-Lain"
                            collector_result[collector_group][song_id] += jumlah
                            collector_all_data["ALL"][song_id] += jumlah

                        except ValueError:
                            print(f"[WARN] Baris {index+1} di file {file}: jumlah tidak valid → {jumlah_cell}")
                            
                    except Exception as e:
                        print(f"[ERROR] Baris {index+1} di file {file}: {str(e)}")
                        continue

            except Exception as e:
                self.log_message(f"Error utama saat proses file {file}: {str(e)}")

        # Simpan output per group
        for group, items in result.items():
            lang_data = defaultdict(list)

            for song_id_raw, jumlah in items.items():
                # 1️⃣ Normalisasi SongId
                song_id_clean = re.sub(r"[A-Za-z]", "", song_id_raw)  # Hapus huruf
                song_id_clean = song_id_clean.lstrip("0")             # Hapus prefix 0

                # 2️⃣ Kategorisasi berdasarkan awal angka
                lang = self.get_language_category(song_id_clean)

                # 3️⃣ Cari kecocokan master SongId (LIKE match)
                matched_info = {"Song": "","RomanSong": "", "Singer": "","SingId": ""}
                final_song_id = song_id_clean  # default
                for master_id, info in song_dict.items():
                    if master_id and song_id_clean in master_id:
                        matched_info = info
                        if master_id != song_id_clean:
                            final_song_id = master_id  # ganti jika tidak persis sama
                        break
                # 4️⃣ Tambahkan ke kategori lang
                if lang in ["Mandarin", "Korea", "Jepang", "Lain-Lain"]:

                    penyanyi_ids = matched_info["SingId"].split(" - ")  # split ID dari "SingerId"
                    penyanyi_ids = [pid.replace(".0", "") for pid in penyanyi_ids]
                    penyanyi_roman = []

                    for pid in penyanyi_ids:
                        roman = sing_dict.get(pid, {}).get("RomanSing", "")  # cari dari sing_dict
                        if roman:
                            penyanyi_roman.append(roman)
                        else:
                            penyanyi_roman.append(pid)  # fallback ke ID jika nama tidak ditemukan

                    lang_data[lang].append({
                        "Judul Lagu": matched_info["RomanSong"],
                        "Penyanyi": " - ".join(penyanyi_roman),
                        "Jumlah Pengguna": jumlah,
                        "ID": final_song_id
                    })
                else:
                    lang_data[lang].append({
                        "Judul Lagu": matched_info["Song"],
                        "Penyanyi": matched_info["Singer"],
                        "Jumlah Pengguna": jumlah,
                        "ID": final_song_id
                    })

            # Buat Excel writer untuk setiap group
            output_file = os.path.join(output_path, f"IDLAGU_Outlet_{group}.xlsx")

            language_categories = {
                "Indonesia Pop": "Indonesia Pop",
                "Indonesia Daerah": "Indonesia Daerah",
                "English": "English",
                "Mandarin": "Mandarin",
                "Jepang": "Jepang",
                "Korea": "Korea",
                "Lain-Lain": "Lain-Lain"
            }

            category_sums = {}

            with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                for lang, records in lang_data.items():
                    df_sheet = pd.DataFrame(records)
                    df_sheet.sort_values(by="Jumlah Pengguna", ascending=False, inplace=True)
                    sheet_name = lang[:31]  # Excel sheet name max length = 31
                    df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

                    worksheet = writer.sheets[sheet_name]
                    row_count = len(df_sheet) + 2  # +2 because of header

                    # Write "Jumlah" and SUM formula in each sheet
                    worksheet.cell(row=row_count, column=1).value = "Jumlah"
                    worksheet.cell(row=row_count, column=3).value = f"=SUM(C2:C{row_count-1})"

                    # Simpan referensi formula hasil jumlah untuk sheet "Total"
                    category_sums[lang] = f"='{sheet_name}'!C{row_count}"

                # Setelah semua sheet selesai, tulis sheet "Total"
                total_sheet = writer.book.create_sheet("Total")

                headers = list(language_categories.keys())
                total_col_letter = get_column_letter(len(headers))

                # 🔷 Merge & judul bagian pertama
                total_sheet.merge_cells('A1:H1')
                total_sheet["A1"].value = "Kategori Bahasa"
                total_sheet["A1"].font = Font(bold=True)
                total_sheet["A1"].alignment = Alignment(horizontal="center")
                total_sheet["A1"].fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")

                # 🔷 Header pertama di A2
                for idx, header in enumerate(headers):
                    cell = total_sheet.cell(row=2, column=idx + 1)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
                    cell.fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")

                # 🔷 Total kolom ke-(len+1)
                total_sheet.cell(row=2, column=len(headers) + 1).value = "Total"
                total_sheet.cell(row=2, column=len(headers) + 1).font = Font(bold=True)
                total_sheet.cell(row=2, column=len(headers) + 1).alignment = Alignment(horizontal="center")
                total_sheet.cell(row=2, column=len(headers) + 1).fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")

                # 🔷 Baris jumlah (dengan formula per sheet) di A3
                sum_row = []
                for lang in headers:
                    formula = category_sums.get(lang, "0")
                    sum_row.append(f"={formula}")
                sum_row.append(f"=SUM(A3:{total_col_letter}3)")
                total_sheet.append(sum_row)

                # 🔷 Merge & judul bagian kedua (A5:H5)
                total_sheet.merge_cells("A5:H5")
                total_sheet["A5"].value = "Persentase"
                total_sheet["A5"].font = Font(bold=True)
                total_sheet["A5"].alignment = Alignment(horizontal="center")
                total_sheet["A5"].fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")

                # 🔷 Header ulang di A6
                for idx, header in enumerate(headers):
                    cell = total_sheet.cell(row=6, column=idx + 1)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
                    cell.fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")

                # 🔷 Header "Total" kolom terakhir
                cell = total_sheet.cell(row=6, column=len(headers) + 1)
                cell.value = "Total"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")

                # 🔷 Baris Persentase di A7
                percent_row = []
                for i in range(len(headers)):
                    col_letter = get_column_letter(i + 1)
                    percent_row.append(f"={col_letter}3/H3")  # Total di kolom terakhir baris 3
                percent_row.append("=H3/H3")
                total_sheet.append(percent_row)

                # Optional: rapikan align center
                for row in total_sheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=8):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center")

            self.log_message(f"File berhasil dibuat: {output_file}")
            if progress_update_callback:
                progress_update_callback()

        # 🔽 Simpan file rekap semua grup
        if all_data["ALL"]:
            lang_data_all = defaultdict(list)
            for song_id_raw, jumlah in all_data["ALL"].items():
                # 🔁 Normalisasi
                song_id_clean = re.sub(r"[A-Za-z]", "", song_id_raw).lstrip("0")

                # 🔁 Kategorisasi
                lang = self.get_language_category(song_id_clean)

                # 🔁 Match master
                matched_info = {"Song": "","RomanSong": "", "Singer": "","SingId": ""}
                final_song_id = song_id_clean
                for master_id, info in song_dict.items():
                    if master_id and song_id_clean in master_id:
                        matched_info = info
                        if master_id != song_id_clean:
                            final_song_id = master_id
                        break

                # 4️⃣ Tambahkan ke kategori lang
                if lang in ["Mandarin", "Korea", "Jepang", "Lain-Lain"]:
                    penyanyi_ids = matched_info["SingId"].split(" - ")  # split ID dari "SingerId"
                    penyanyi_ids = [pid.replace(".0", "") for pid in penyanyi_ids]
                    penyanyi_roman = []

                    for pid in penyanyi_ids:
                        roman = sing_dict.get(pid, {}).get("RomanSing", "")  # cari dari sing_dict
                        if roman:
                            penyanyi_roman.append(roman)
                        else:
                            penyanyi_roman.append(pid)  # fallback ke ID jika nama tidak ditemukan

                    lang_data_all[lang].append({
                        "Judul Lagu": matched_info["RomanSong"],
                        "Penyanyi": " - ".join(penyanyi_roman),
                        "Jumlah Pengguna": jumlah,
                        "ID": final_song_id
                    })
                else:
                    lang_data_all[lang].append({
                        "Judul Lagu": matched_info["Song"],
                        "Penyanyi": matched_info["Singer"],
                        "Jumlah Pengguna": jumlah,
                        "ID": final_song_id
                    })

            # 🔁 Simpan ke file
            output_file_all = os.path.join(output_path, "IDLAGU_ALL.xlsx")

            language_categories = {
                "Indonesia Pop": "Indonesia Pop",
                "Indonesia Daerah": "Indonesia Daerah",
                "English": "English",
                "Mandarin": "Mandarin",
                "Jepang": "Jepang",
                "Korea": "Korea",
                "Lain-Lain": "Lain-Lain"
            }

            category_sums = {}

            with pd.ExcelWriter(output_file_all, engine="openpyxl") as writer:
                for lang, records in lang_data_all.items():
                    df_sheet = pd.DataFrame(records)
                    df_sheet.sort_values(by="Jumlah Pengguna", ascending=False, inplace=True)
                    sheet_name = lang[:31]  # Excel sheet name max length = 31
                    df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

                    worksheet = writer.sheets[sheet_name]
                    row_count = len(df_sheet) + 2  # +2 because of header

                    # Write "Jumlah" and SUM formula in each sheet
                    worksheet.cell(row=row_count, column=1).value = "Jumlah"
                    worksheet.cell(row=row_count, column=3).value = f"=SUM(C2:C{row_count-1})"

                    # Simpan referensi formula hasil jumlah untuk sheet "Total"
                    category_sums[lang] = f"='{sheet_name}'!C{row_count}"

                # Setelah semua sheet selesai, tulis sheet "Total"
                total_sheet = writer.book.create_sheet("Total")

                headers = list(language_categories.keys())
                total_col_letter = get_column_letter(len(headers))

                # 🔷 Merge & judul bagian pertama
                total_sheet.merge_cells('A1:H1')
                total_sheet["A1"].value = "Kategori Bahasa"
                total_sheet["A1"].font = Font(bold=True)
                total_sheet["A1"].alignment = Alignment(horizontal="center")
                total_sheet["A1"].fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")

                # 🔷 Header pertama di A2
                for idx, header in enumerate(headers):
                    cell = total_sheet.cell(row=2, column=idx + 1)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
                    cell.fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")

                # 🔷 Total kolom ke-(len+1)
                total_sheet.cell(row=2, column=len(headers) + 1).value = "Total"
                total_sheet.cell(row=2, column=len(headers) + 1).font = Font(bold=True)
                total_sheet.cell(row=2, column=len(headers) + 1).alignment = Alignment(horizontal="center")
                total_sheet.cell(row=2, column=len(headers) + 1).fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")

                # 🔷 Baris jumlah (dengan formula per sheet) di A3
                sum_row = []
                for lang in headers:
                    formula = category_sums.get(lang, "0")
                    sum_row.append(f"={formula}")
                sum_row.append(f"=SUM(A3:{total_col_letter}3)")
                total_sheet.append(sum_row)

                # 🔷 Merge & judul bagian kedua (A5:H5)
                total_sheet.merge_cells("A5:H5")
                total_sheet["A5"].value = "Persentase"
                total_sheet["A5"].font = Font(bold=True)
                total_sheet["A5"].alignment = Alignment(horizontal="center")
                total_sheet["A5"].fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")

                # 🔷 Header ulang di A6
                for idx, header in enumerate(headers):
                    cell = total_sheet.cell(row=6, column=idx + 1)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
                    cell.fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")

                # 🔷 Header "Total" kolom terakhir
                cell = total_sheet.cell(row=6, column=len(headers) + 1)
                cell.value = "Total"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")

                # 🔷 Baris Persentase di A7
                percent_row = []
                for i in range(len(headers)):
                    col_letter = get_column_letter(i + 1)
                    percent_row.append(f"={col_letter}3/H3")  # Total di kolom terakhir baris 3
                percent_row.append("=H3/H3")
                total_sheet.append(percent_row)

                # Optional: rapikan align center
                for row in total_sheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=8):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center")

            self.log_message(f"File gabungan berhasil dibuat: {output_file_all}")
            if progress_update_callback:
                progress_update_callback()
    
            collector_categories = {
                "KCI": "KCI",
                "WAMI": "WAMI", 
                "RAI": "RAI",
                "Lain-Lain": "Lain-Lain"
            }

        for group, items in result.items():
            collector_data = defaultdict(list)
            
            for song_id_raw, jumlah in items.items():
                song_id_clean = re.sub(r"[A-Za-z]", "", song_id_raw).lstrip("0")
                
                # Match dengan master data
                matched_info = {"Song": "","RomanSong": "", "Singer": "","SingId": "", "Label": ""}
                final_song_id = song_id_clean  # default
                
                for master_id, info in song_dict.items():
                    if master_id and song_id_clean in master_id:
                        matched_info = info
                        if master_id != song_id_clean:
                            final_song_id = master_id
                        break
                
                label = matched_info.get("Label", "").upper()
                collector_group = label if label in ["KCI", "WAMI", "RAI"] else "Lain-Lain"
                
                # Tentukan Roman Song/Singer berdasarkan bahasa
                lang = self.get_language_category(song_id_clean)
                
                if lang in ["Mandarin", "Korea", "Jepang", "Lain-Lain"]:
                    # Gunakan Roman Song dan Roman Singer
                    penyanyi_ids = matched_info.get("SingId", "").split(" - ")
                    penyanyi_roman = [sing_dict.get(pid, {}).get("RomanSing", pid) for pid in penyanyi_ids]
                    
                    collector_data[collector_group].append({
                        "Judul Lagu": matched_info.get("RomanSong", song_id_clean),
                        "Penyanyi": " - ".join(penyanyi_roman),
                        "Jumlah Pengguna": jumlah,
                        "ID": song_id_raw
                    })
                else:
                    # Gunakan original Song dan Singer
                    collector_data[collector_group].append({
                        "Judul Lagu": matched_info.get("Song", song_id_clean),
                        "Penyanyi": matched_info.get("Singer", ""),
                        "Jumlah Pengguna": jumlah,
                        "ID": song_id_raw
                    })

            # Simpan file outlet dengan sheet by collector
            output_file = os.path.join(output_path, f"LABEL_Outlet_{group}.xlsx")
            
            category_sums_collector = {}
            
            with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                for collector, records in collector_data.items():
                    df_sheet = pd.DataFrame(records)
                    df_sheet.sort_values(by="Jumlah Pengguna", ascending=False, inplace=True)
                    sheet_name = collector[:31]
                    df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Tambahkan total
                    worksheet = writer.sheets[sheet_name]
                    row_count = len(df_sheet) + 2
                    worksheet.cell(row=row_count, column=1).value = "Jumlah"
                    worksheet.cell(row=row_count, column=3).value = f"=SUM(C2:C{row_count-1})"
                    
                    # Simpan referensi formula untuk sheet Total
                    category_sums_collector[collector] = f"='{sheet_name}'!C{row_count}"
                
                # 🆕 TAMBAH SHEET TOTAL UNTUK COLLECTOR
                total_sheet = writer.book.create_sheet("Total")
                
                headers_collector = list(collector_categories.keys())
                total_col_letter = get_column_letter(len(headers_collector))
                
                # 🔷 Merge & judul bagian pertama
                total_sheet.merge_cells('A1:E1')
                total_sheet["A1"].value = "Kategori Label"
                total_sheet["A1"].font = Font(bold=True)
                total_sheet["A1"].alignment = Alignment(horizontal="center")
                total_sheet["A1"].fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")
                
                # 🔷 Header pertama di A2
                for idx, header in enumerate(headers_collector):
                    cell = total_sheet.cell(row=2, column=idx + 1)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
                    cell.fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")
                
                # 🔷 Total kolom terakhir
                total_sheet.cell(row=2, column=len(headers_collector) + 1).value = "Total"
                total_sheet.cell(row=2, column=len(headers_collector) + 1).font = Font(bold=True)
                total_sheet.cell(row=2, column=len(headers_collector) + 1).alignment = Alignment(horizontal="center")
                total_sheet.cell(row=2, column=len(headers_collector) + 1).fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")
                
                # 🔷 Baris jumlah (dengan formula per sheet) di A3
                sum_row = []
                for collector in headers_collector:
                    formula = category_sums_collector.get(collector, "0")
                    sum_row.append(f"={formula}")
                sum_row.append(f"=SUM(A3:{total_col_letter}3)")
                total_sheet.append(sum_row)
                
                # 🔷 Merge & judul bagian kedua (A5:E5)
                total_sheet.merge_cells("A5:E5")
                total_sheet["A5"].value = "Persentase"
                total_sheet["A5"].font = Font(bold=True)
                total_sheet["A5"].alignment = Alignment(horizontal="center")
                total_sheet["A5"].fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")
                
                # 🔷 Header ulang di A6
                for idx, header in enumerate(headers_collector):
                    cell = total_sheet.cell(row=6, column=idx + 1)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
                    cell.fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")
                
                # 🔷 Header "Total" kolom terakhir
                cell = total_sheet.cell(row=6, column=len(headers_collector) + 1)
                cell.value = "Total"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")
                
                # 🔷 Baris Persentase di A7
                percent_row = []
                total_col_letter_percent = get_column_letter(len(headers_collector) + 1)
                for i in range(len(headers_collector)):
                    col_letter = get_column_letter(i + 1)
                    percent_row.append(f"={col_letter}3/{total_col_letter_percent}3")
                percent_row.append(f"={total_col_letter_percent}3/{total_col_letter_percent}3")
                total_sheet.append(percent_row)
                
                # Optional: rapikan align center
                for row in total_sheet.iter_rows(min_row=1, max_row=7, min_col=1, max_col=len(headers_collector)+1):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center")
            
            self.log_message(f"File LABEL berhasil dibuat: {output_file}")
            if progress_update_callback:
                progress_update_callback()

        # 🆕 BUAT FILE ALL_DATA UNTUK COLLECTOR
        if collector_all_data["ALL"]:
            collector_data_all = defaultdict(list)
            
            for song_id_raw, jumlah in collector_all_data["ALL"].items():
                song_id_clean = re.sub(r"[A-Za-z]", "", song_id_raw).lstrip("0")
                
                # Match dengan master data
                matched_info = {"Song": "","RomanSong": "", "Singer": "","SingId": "", "Label": ""}
                final_song_id = song_id_clean  # default
                
                for master_id, info in song_dict.items():
                    if master_id and song_id_clean in master_id:
                        matched_info = info
                        if master_id != song_id_clean:
                            final_song_id = master_id  # ganti jika tidak persis sama
                        break
                
                label = matched_info.get("Label", "").upper()
                collector_group = label if label in ["KCI", "WAMI", "RAI"] else "Lain-Lain"
                
                # Tentukan Roman Song/Singer berdasarkan bahasa
                lang = self.get_language_category(song_id_clean)
                
                if lang in ["Mandarin", "Korea", "Jepang", "Lain-Lain"]:
                    # Gunakan Roman Song dan Roman Singer
                    penyanyi_ids = matched_info.get("SingId", "").split(" - ")
                    penyanyi_roman = [sing_dict.get(pid, {}).get("RomanSing", pid) for pid in penyanyi_ids]
                    
                    collector_data_all[collector_group].append({
                        "Judul Lagu": matched_info.get("RomanSong", song_id_clean),
                        "Penyanyi": " - ".join(penyanyi_roman),
                        "Jumlah Pengguna": jumlah,
                        "ID": song_id_raw
                    })
                else:
                    # Gunakan original Song dan Singer
                    collector_data_all[collector_group].append({
                        "Judul Lagu": matched_info.get("Song", song_id_clean),
                        "Penyanyi": matched_info.get("Singer", ""),
                        "Jumlah Pengguna": jumlah,
                        "ID": song_id_raw
                    })
            
            # Simpan ke file ALL_DATA collector
            output_file_collector_all = os.path.join(output_path, "LABEL_ALL.xlsx")
            
            category_sums_collector_all = {}
            
            with pd.ExcelWriter(output_file_collector_all, engine="openpyxl") as writer:
                for collector, records in collector_data_all.items():
                    df_sheet = pd.DataFrame(records)
                    df_sheet.sort_values(by="Jumlah Pengguna", ascending=False, inplace=True)
                    sheet_name = collector[:31]
                    df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Tambahkan total
                    worksheet = writer.sheets[sheet_name]
                    row_count = len(df_sheet) + 2
                    worksheet.cell(row=row_count, column=1).value = "Jumlah"
                    worksheet.cell(row=row_count, column=3).value = f"=SUM(C2:C{row_count-1})"
                    
                    # Simpan referensi formula untuk sheet Total
                    category_sums_collector_all[collector] = f"='{sheet_name}'!C{row_count}"
                
                # SHEET TOTAL UNTUK COLLECTOR ALL_DATA
                total_sheet = writer.book.create_sheet("Total")
                
                headers_collector = list(collector_categories.keys())
                total_col_letter = get_column_letter(len(headers_collector))
                
                # 🔷 Merge & judul bagian pertama
                total_sheet.merge_cells('A1:E1')
                total_sheet["A1"].value = "Kategori Label"
                total_sheet["A1"].font = Font(bold=True)
                total_sheet["A1"].alignment = Alignment(horizontal="center")
                total_sheet["A1"].fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")
                
                # 🔷 Header pertama di A2
                for idx, header in enumerate(headers_collector):
                    cell = total_sheet.cell(row=2, column=idx + 1)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
                    cell.fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")
                
                # 🔷 Total kolom terakhir
                total_sheet.cell(row=2, column=len(headers_collector) + 1).value = "Total"
                total_sheet.cell(row=2, column=len(headers_collector) + 1).font = Font(bold=True)
                total_sheet.cell(row=2, column=len(headers_collector) + 1).alignment = Alignment(horizontal="center")
                total_sheet.cell(row=2, column=len(headers_collector) + 1).fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")
                
                # 🔷 Baris jumlah (dengan formula per sheet) di A3
                sum_row = []
                for collector in headers_collector:
                    formula = category_sums_collector_all.get(collector, "0")
                    sum_row.append(f"={formula}")
                sum_row.append(f"=SUM(A3:{total_col_letter}3)")
                total_sheet.append(sum_row)
                
                # 🔷 Merge & judul bagian kedua (A5:E5)
                total_sheet.merge_cells("A5:E5")
                total_sheet["A5"].value = "Persentase"
                total_sheet["A5"].font = Font(bold=True)
                total_sheet["A5"].alignment = Alignment(horizontal="center")
                total_sheet["A5"].fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")
                
                # 🔷 Header ulang di A6
                for idx, header in enumerate(headers_collector):
                    cell = total_sheet.cell(row=6, column=idx + 1)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
                    cell.fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")
                
                # 🔷 Header "Total" kolom terakhir
                cell = total_sheet.cell(row=6, column=len(headers_collector) + 1)
                cell.value = "Total"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid")
                
                # 🔷 Baris Persentase di A7
                percent_row = []
                total_col_letter_percent = get_column_letter(len(headers_collector) + 1)
                for i in range(len(headers_collector)):
                    col_letter = get_column_letter(i + 1)
                    percent_row.append(f"={col_letter}3/{total_col_letter_percent}3")
                percent_row.append(f"={total_col_letter_percent}3/{total_col_letter_percent}3")
                total_sheet.append(percent_row)
                
                # Optional: rapikan align center
                for row in total_sheet.iter_rows(min_row=1, max_row=7, min_col=1, max_col=len(headers_collector)+1):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center")
            
            self.log_message(f"File LABEL ALL berhasil dibuat: {output_file_collector_all}")
            if progress_update_callback:
                progress_update_callback()
                
    def get_language_category(self, song_id):
        if song_id.startswith(("10", "11", "12", "13", "14", "15", "16", "17", "19")):
            return "Indonesia Pop"
        elif song_id.startswith("18"):
            return "Indonesia Daerah"
        elif song_id.startswith("2"):
            return "English"
        elif song_id.startswith("3"):
            return "Mandarin"
        elif song_id.startswith("4"):
            return "Jepang"
        elif song_id.startswith("5"):
            return "Korea"
        return "Lain-Lain"

    def update_progress(self):
        self.progress["value"] += 1
        self.root.update_idletasks()

    def process_txt_to_xls(self, txt_file, xls_file):
        try:
            # Read the TXT file line by line
            with open(txt_file, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            
            # Process each line to extract first and fourth columns
            data = []
            for line in lines:
                parts = line.strip().split('||')
                if len(parts) >= 4:  # Ensure we have at least 4 columns
                    song_id = parts[0].strip()
                    count = parts[3].strip()
                    
                    # Only add if both values exist and count is numeric
                    if song_id and count and count.isdigit():
                        data.append([song_id, int(count)])
            
            # Create a DataFrame and group by song ID to sum counts
            df = pd.DataFrame(data, columns=['ID', 'Jumlah'])
            df = df.groupby('ID', as_index=False)['Jumlah'].sum()
            
            # Sort by count in descending order
            df = df.sort_values('Jumlah', ascending=False)
            
            # Save to XLS file
            df.to_excel(xls_file, index=False,sheet_name='Lap1')
            
            return True
        
        except Exception as e:
            self.log_message(f"Error processing {txt_file}: {str(e)}")
            return False


if __name__ == "__main__":
    root = tk.Tk()
    app = GenerateTopHit(root)
    root.mainloop()
